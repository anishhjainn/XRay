# processors/xlsx_processor.py
"""
XLSX technician (FileProcessor): builds a read-only FileArtifact for .xlsx files.

This module extracts facts (read-only metadata). Checks consume these facts and
decide pass/fail. Keeping processors and checks separate follows SRP and DIP.

Metadata provided (keys in artifact.metadata):

Identity & core props
- kind: "xlsx"
- core_author: str | None
- core_created: ISO 8601 | None
- core_modified: ISO 8601 | None

Workbook structure & visibility
- sheet_count: int
- hidden_sheet_count: int
- very_hidden_sheet_count: int

Formulas & errors
- formula_count: int                  (# of <f> tags across all worksheets)
- error_cell_count: int               (# of <c t="e"> cells – error-typed cells)
- formula_ref_error_count: int        (# of formulas containing '#REF!')
- other_error_token_count: int        (# of other error tokens found in formulas/values)

Comments (legacy notes) & threaded comments
- comments_count: int                 (legacy notes in xl/comments*.xml)
- threaded_comments_count: int        (modern threaded comments in xl/threadedComments/*.xml)

Externalities & data connections
- external_links_count: int           (rels with TargetMode="External" or URL/UNC targets)
- data_connections_count: int         (# of <connection> in xl/connections.xml)

Protection & macros
- workbook_structure_protected: bool  (structure/windows protection present)
- password_encrypted_workbook: bool   (file-level encryption: not a plain OOXML ZIP)
- has_vba_project: bool               (xl/vbaProject.bin present; unexpected for .xlsx)

Yellow highlighting (cells and sheet tabs)
- yellow_cell_count: int
- yellow_tab_sheets: list[str]
- yellow_tab_sheet_count: int

Hidden rows/columns (reserved for future UI; not fully populated here)
- hidden_row_count: int
- hidden_col_count: int
- hidden_rows_sample: list
- hidden_cols_sample: list

Diagnostics
- read_error: bool
- read_error_detail: str | None

New (Phase 3: for spelling now, grammar later)
- text_sample: str                    (normalized plain text, up to max_text_chars)
- text_length: int
- token_count: int
- text_extraction_error: bool
- text_extraction_error_detail: str | None
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, Optional
from zipfile import ZipFile, BadZipFile
from xml.etree.ElementTree import iterparse

# Third-party: used minimally, in read-only mode for core props & sheet states
from openpyxl import load_workbook

from services.xlsx_theme import resolve_theme_color, theme_rgb_map_from_path

from core.interfaces import FileProcessor
from core.models import FileArtifact
from core.registry import register_processor

# NEW: spelling/grammar text extraction helpers and config
from utils.text_extract import extract_xlsx_text, tokenize_words
from infra.config_loader import load_config


# Tokens that commonly indicate Excel calculation errors. We count occurrences found
# in formula text or value text where accessible (defensive; t="e" cells already counted).
_ERROR_TOKENS = {
    "#DIV/0!",
    "#NAME?",
    "#VALUE!",
    "#NUM!",
    "#N/A",
    "#NULL!",
    "#GETTING_DATA",
}


def _norm_rgb(rgb: str | None) -> str:
    """Normalize ARGB/RGB to last 6 hex chars; e.g. '00FFFF00' -> 'FFFF00'."""
    s = (rgb or "").upper()
    return s[-6:] if len(s) >= 6 else s


def _color_obj_is_classic_yellow(color, theme_rgb_map: dict[int, str]) -> bool:
    """
    Decide whether an openpyxl color object represents classic Excel yellow.

    Classic yellow detection:
    - Solid pattern with fg/bg rgb 'FFFF00'
    - indexed == 6
    - theme+tint resolves to 'FFFF00'
    """
    if color is None:
        return False

    rgb_value = getattr(color, "rgb", None)
    if isinstance(rgb_value, str) and _norm_rgb(rgb_value) == "FFFF00":
        return True

    idx = getattr(color, "indexed", None)
    if idx is not None and str(idx) == "6":
        return True

    theme_idx = getattr(color, "theme", None)
    if theme_idx is not None:
        resolved = resolve_theme_color(theme_rgb_map, theme_idx, getattr(color, "tint", None))
        if resolved and _norm_rgb(resolved) == "FFFF00":
            return True

    return False


def _is_classic_yellow_fill(fill, theme_rgb_map: dict[int, str]) -> bool:
    """Return True if the cell fill is a 'solid' classic yellow (fg/bg)."""
    try:
        if getattr(fill, "patternType", None) != "solid":
            return False
        fg = getattr(fill, "fgColor", None)
        bg = getattr(fill, "bgColor", None)
        if _color_obj_is_classic_yellow(fg, theme_rgb_map):
            return True
        if _color_obj_is_classic_yellow(bg, theme_rgb_map):
            return True
        return False
    except Exception:
        return False


class XlsxProcessor(FileProcessor):
    def supports(self):
        # LSP: mirrors other processors
        return [".xlsx"]

    def build_artifact(self, path: Path) -> FileArtifact:
        size = path.stat().st_size

        # Default metadata. Populate defensively; keep the artifact useful even if some reads fail.
        metadata: Dict[str, Any] = {
            "kind": "xlsx",
            "core_author": None,
            "core_created": None,
            "core_modified": None,
            "sheet_count": None,
            "hidden_sheet_count": 0,
            "very_hidden_sheet_count": 0,
            "formula_count": 0,
            "error_cell_count": 0,
            "formula_ref_error_count": 0,
            "other_error_token_count": 0,
            "comments_count": 0,
            "threaded_comments_count": 0,
            "external_links_count": 0,
            "data_connections_count": 0,
            "workbook_structure_protected": False,
            "password_encrypted_workbook": False,
            "has_vba_project": False,
            "read_error": False,
            "yellow_cell_count": 0,
            "yellow_tab_sheets": [],
            "yellow_tab_sheet_count": 0,
            "hidden_row_count": 0,
            "hidden_col_count": 0,
            "hidden_rows_sample": [],   # list of {"sheet": <name>, "row": <int>}
            "hidden_cols_sample": [],   # list of {"sheet": <name>, "col": <str>}
            # NEW: text extraction for spelling/grammar checks
            "text_sample": "",
            "text_length": 0,
            "token_count": 0,
            "text_extraction_error": False,
            # "text_extraction_error_detail": "...",
        }

        theme_rgb_map = theme_rgb_map_from_path(path)

        # --- 1) Quick header sniff: distinguish OOXML ZIP vs OLE/encrypted container ---
        # Password-encrypted Excel typically uses OLE CF, not a plain ZIP. Plain XLSX starts with PK 0x03 0x04.
        try:
            with path.open("rb") as fp:
                magic = fp.read(8)
            if magic.startswith(b"\xD0\xCF\x11\xE0"):  # OLE Compound (encrypted or legacy)
                metadata["password_encrypted_workbook"] = True
                metadata["read_error"] = True
                metadata["read_error_detail"] = "Encrypted or non-OOXML container (OLE compound)"
                # Early return: we can't parse styles/parts from an encrypted OLE container
                return FileArtifact(
                    path=path,
                    extension=".xlsx",
                    size_bytes=size,
                    metadata=metadata,
                )
        except Exception as exc:
            # We can still try openpyxl/zip later; record a hint
            metadata["read_error"] = True
            metadata["read_error_detail"] = f"Header read failed: {exc.__class__.__name__}"

        # --- 2) Use openpyxl for core props, sheet visibility, yellow detection ---
        try:
            # Use data_only=True to resolve cached values where possible.
            # Note: We are NOT in read_only mode here because we need styles for fill/tab color inspection.
            wb = load_workbook(filename=str(path), data_only=True)

            # Core document properties (author/created/modified)
            props = getattr(wb, "properties", None)
            if props:
                metadata["core_author"] = getattr(props, "creator", None) or getattr(props, "lastModifiedBy", None)
                metadata["core_created"] = _safe_iso(getattr(props, "created", None))
                metadata["core_modified"] = _safe_iso(getattr(props, "modified", None))

            # Sheet list & visibility states
            sheetnames = list(getattr(wb, "sheetnames", []) or [])
            metadata["sheet_count"] = len(sheetnames)

            hidden = 0
            very_hidden = 0
            for name in sheetnames:
                try:
                    ws = wb[name]
                    state = getattr(ws, "sheet_state", "visible") or "visible"
                    if state == "hidden":
                        hidden += 1
                    elif state == "veryHidden":
                        very_hidden += 1
                except Exception:
                    # Skip problematic sheets but keep scanning
                    continue
            metadata["hidden_sheet_count"] = hidden
            metadata["very_hidden_sheet_count"] = very_hidden

            # ---- Yellow sheet tabs (classic yellow) ----
            yellow_tabs = []
            for name in sheetnames:
                try:
                    ws = wb[name]
                    tab = getattr(getattr(ws, "sheet_properties", None), "tabColor", None)
                    if tab is not None and _color_obj_is_classic_yellow(tab, theme_rgb_map):
                        yellow_tabs.append(name)
                except Exception:
                    continue
            metadata["yellow_tab_sheets"] = yellow_tabs
            metadata["yellow_tab_sheet_count"] = len(yellow_tabs)

            # ---- Yellow cells (classic yellow) ----
            yellow_cells = 0
            try:
                for ws in wb.worksheets:
                    for row in ws.iter_rows():
                        for cell in row:
                            fill = getattr(cell, "fill", None)
                            if fill and _is_classic_yellow_fill(fill, theme_rgb_map):
                                yellow_cells += 1
            except Exception as exc:
                # Continue with partial counts; record a warning detail
                metadata["read_error"] = True
                prev = metadata.get("read_error_detail")
                msg = f"yellow-cells-scan: {exc.__class__.__name__}"
                metadata["read_error_detail"] = msg if not prev else f"{prev}; {msg}"
            metadata["yellow_cell_count"] = yellow_cells

        except Exception as exc:
            # Not fatal: XML streaming (below) can still produce most counts
            metadata.setdefault("sheet_count", None)
            metadata["read_error"] = True
            detail = str(exc) or exc.__class__.__name__
            prev = metadata.get("read_error_detail")
            metadata["read_error_detail"] = detail if not prev else f"{prev}; openpyxl: {detail}"

        # --- 3) ZIP + streaming XML: formulas/errors/comments/external-links/connections/protection/VBA ---
        try:
            with ZipFile(str(path)) as zf:
                names = set(zf.namelist())

                # VBA presence: unexpected in .xlsx (should be .xlsm)
                if "xl/vbaProject.bin" in names:
                    metadata["has_vba_project"] = True

                # Workbook protection via <workbookProtection>
                if "xl/workbook.xml" in names:
                    metadata["workbook_structure_protected"] = _workbook_structure_is_protected(zf, "xl/workbook.xml")

                # External links via workbook relationships
                if "xl/_rels/workbook.xml.rels" in names:
                    ext_count = _count_external_relationships(zf, "xl/_rels/workbook.xml.rels")
                    metadata["external_links_count"] += ext_count

                # externalLinks parts also indicate external references
                for n in names:
                    if n.startswith("xl/externalLinks/") and n.endswith(".xml"):
                        metadata["external_links_count"] += 1

                # Data connections
                if "xl/connections.xml" in names:
                    metadata["data_connections_count"] = _count_tag_local(zf, "xl/connections.xml", "connection")

                # Comments (legacy)
                for n in names:
                    if n.startswith("xl/comments") and n.endswith(".xml"):
                        metadata["comments_count"] += _count_tag_local(zf, n, "comment")

                # Threaded comments (modern)
                for n in names:
                    if n.startswith("xl/threadedComments/") and n.endswith(".xml"):
                        metadata["threaded_comments_count"] += _count_tag_local(zf, n, "threadedComment")

                # Worksheet scans for formulas and error cells/tokens
                for n in names:
                    if n.startswith("xl/worksheets/sheet") and n.endswith(".xml"):
                        f_count, e_count, ref_err, other_errs = _scan_worksheet_for_formulas_and_errors(zf, n)
                        metadata["formula_count"] += f_count
                        metadata["error_cell_count"] += e_count
                        metadata["formula_ref_error_count"] += ref_err
                        metadata["other_error_token_count"] += other_errs

        except (BadZipFile, OSError, ValueError, KeyError, Exception) as exc:
            metadata["read_error"] = True
            detail = str(exc) or exc.__class__.__name__
            prev = metadata.get("read_error_detail")
            metadata["read_error_detail"] = detail if not prev else f"{prev}; zip-scan: {detail}"

        # --- 4) NEW: Plain-text sample extraction for spelling/grammar checks ---
        # We keep this separate from the structural scans above and do not flip read_error on extraction failure.
        try:
            cfg = load_config()
            if cfg.get("enable_spelling", True) or cfg.get("enable_grammar", False):
                max_chars = int(cfg.get("max_text_chars", 5_000_000))
                # Your policy: include text from hidden sheets; skip formula texts entirely.
                sample = extract_xlsx_text(
                    path,
                    max_chars=max_chars,
                    include_hidden=True,
                    skip_formulas=True,
                )
                metadata["text_sample"] = sample
                metadata["text_length"] = len(sample)
                metadata["token_count"] = len(tokenize_words(sample))
        except Exception as exc:
            metadata["text_sample"] = ""
            metadata["text_length"] = 0
            metadata["token_count"] = 0
            metadata["text_extraction_error"] = True
            metadata["text_extraction_error_detail"] = str(exc) or exc.__class__.__name__

        # Return the immutable artifact the rest of the system expects
        return FileArtifact(
            path=path,
            extension=".xlsx",
            size_bytes=size,
            metadata=metadata,
        )


# --- helpers (module-internal) -------------------------------------------------


def _safe_iso(dt) -> Optional[str]:
    """Convert a datetime-like object to ISO 8601 if available, else None."""
    try:
        return dt.isoformat() if dt is not None else None
    except Exception:
        return None


def _local_name(tag: str) -> str:
    """Strip the XML namespace from a tag: '{namespace}name' -> 'name'."""
    if "}" in tag:
        return tag.split("}", 1)[1]
    return tag


def _count_tag_local(zf: ZipFile, member: str, local_name: str) -> int:
    """
    Count how many elements with the given local-name appear in an XML part.

    We stream with iterparse(events=("end",)) and clear elements to keep memory
    usage low for large sheets.
    """
    count = 0
    with zf.open(member) as fp:
        for _event, elem in iterparse(fp, events=("end",)):
            if _local_name(elem.tag) == local_name:
                count += 1
            elem.clear()
    return count


def _workbook_structure_is_protected(zf: ZipFile, member: str) -> bool:
    """
    Detect structure/windows protection in xl/workbook.xml via <workbookProtection>.
    We treat presence of the element (with typical attributes) as 'protected'.
    """
    protected = False
    with zf.open(member) as fp:
        for _event, elem in iterparse(fp, events=("end",)):
            if _local_name(elem.tag) == "workbookProtection":
                if elem.attrib:
                    protected = True
                elem.clear()
            else:
                elem.clear()
    return protected


def _looks_external_target(target: str) -> bool:
    """Heuristic to decide if a relationship target is external (URL or UNC path)."""
    t = (target or "").strip().lower()
    return (
        t.startswith("http://")
        or t.startswith("https://")
        or t.startswith("ftp://")
        or t.startswith("\\\\")  # UNC path
    )


def _count_external_relationships(zf: ZipFile, member: str) -> int:
    """
    Count relationships in xl/_rels/workbook.xml.rels that are 'external':
    - TargetMode="External", or
    - Target looks like URL/UNC.
    """
    count = 0
    with zf.open(member) as fp:
        for _event, elem in iterparse(fp, events=("end",)):
            if _local_name(elem.tag) == "Relationship":
                mode = elem.attrib.get("TargetMode", "")
                target = elem.attrib.get("Target", "")
                if mode == "External" or _looks_external_target(target):
                    count += 1
            elem.clear()
    return count


def _scan_worksheet_for_formulas_and_errors(zf: ZipFile, member: str) -> tuple[int, int, int, int]:
    """
    Stream a worksheet XML (xl/worksheets/sheet*.xml) and return:
      (formula_count, error_cell_count, formula_ref_error_count, other_error_token_count)

    We look for:
      - <f>...</f>       -> formulas (formula_count)
      - <c t="e">...</c> -> error-typed cells (error_cell_count)
      - '#REF!' inside formula text -> formula_ref_error_count
      - Other error tokens in formula text or cell value -> other_error_token_count
    """
    formula_count = 0
    error_cells = 0
    ref_err = 0
    other_errs = 0

    with zf.open(member) as fp:
        for _event, elem in iterparse(fp, events=("end",)):
            lname = _local_name(elem.tag)

            if lname == "f":
                formula_count += 1
                text = (elem.text or "").strip()
                if "#REF!" in text:
                    ref_err += 1
                if any(tok in text for tok in _ERROR_TOKENS):
                    other_errs += 1
                elem.clear()

            elif lname == "c":
                cell_type = (elem.attrib.get("t") or "").strip().lower()
                if cell_type == "e":
                    error_cells += 1
                elem.clear()

            elif lname == "v":
                text = (elem.text or "").strip()
                if any(tok == text for tok in _ERROR_TOKENS):
                    other_errs += 1
                elem.clear()

            else:
                elem.clear()

    return formula_count, error_cells, ref_err, other_errs


# Register on import (OCP/DIP: plugins self-register; no central switch/case grows)
register_processor(XlsxProcessor())
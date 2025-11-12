# services/xlsx_locators_openpyxl.py
"""
On-demand XLSX locator using openpyxl (simple & readable).

Public API:
    list_yellow_cells(path: Path) -> list[tuple[str, str]]

Classic Excel yellow detection:
- Solid pattern fill and fg/bg color resolves to RGB 'FFFF00' (normalized), or
- Indexed color 6 (Excel's standard yellow), or
- Theme color (plus tint) that resolves to RGB 'FFFF00'
"""

from __future__ import annotations

from pathlib import Path
from typing import List, Tuple

from openpyxl import load_workbook

from services.xlsx_theme import resolve_theme_color, theme_rgb_map_from_path


def _norm_rgb(rgb: str | None) -> str:
    # Normalize ARGB/RGB to the last 6 hex digits; e.g. "00FFFF00" -> "FFFF00"
    s = (rgb or "").upper()
    return s[-6:] if len(s) >= 6 else s


def _color_obj_is_classic_yellow(color, theme_rgb_map: dict[int, str]) -> bool:
    if color is None:
        return False

    rgb_value = getattr(color, "rgb", None)
    if isinstance(rgb_value, str) and _norm_rgb(rgb_value) == "FFFF00":
        return True

    idx = getattr(color, "indexed", None)
    if idx is not None and str(idx) == "6":
        return True

    theme_idx = getattr(color, "theme", None)
    if theme_idx is not None:
        resolved = resolve_theme_color(theme_rgb_map, theme_idx, getattr(color, "tint", None))
        if resolved and _norm_rgb(resolved) == "FFFF00":
            return True

    return False


def _is_classic_yellow_fill(fill, theme_rgb_map: dict[int, str]) -> bool:
    """
    Return True if the openpyxl cell fill is a 'solid' classic-yellow fill.
    Handles:
      - fgColor.rgb / bgColor.rgb normalized to 'FFFF00'
      - indexed color '6' (Excel's yellow)
      - theme color + tint that resolves to 'FFFF00'
    """
    try:
        if getattr(fill, "patternType", None) != "solid":
            return False

        fg = getattr(fill, "fgColor", None)
        bg = getattr(fill, "bgColor", None)
        if _color_obj_is_classic_yellow(fg, theme_rgb_map):
            return True
        if _color_obj_is_classic_yellow(bg, theme_rgb_map):
            return True
        return False
    except Exception:
        return False


def list_yellow_cells(path: Path) -> List[Tuple[str, str]]:
    """
    Return a list of (sheet_name, cell_ref) where cells are classic-yellow.
    """
    p = Path(path)
    if not p.exists() or not p.is_file():
        raise FileNotFoundError(f"No such file: {p}")

    # Need the full workbook (read_only=False) because ReadOnlyCell doesn't expose style/fill info.
    wb = load_workbook(filename=str(p), data_only=True)
    theme_rgb_map = theme_rgb_map_from_path(p)

    results: List[Tuple[str, str]] = []
    for ws in wb.worksheets:
        # iter_rows keeps memory modest; good enough for typical files
        for row in ws.iter_rows():
            for cell in row:
                # Some read_only cells can be empty stubs; guard carefully
                fill = getattr(cell, "fill", None)
                if fill and _is_classic_yellow_fill(fill, theme_rgb_map):
                    results.append((ws.title, cell.coordinate))
    return results
